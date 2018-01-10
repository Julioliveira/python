#!C:\Users\Julio\AppData\Local\Programs\Python\Python36-32
import openpyxl
from datetime import datetime
from sys import argv as arguments


def axpe():
    nome = arguments[2]
    nome += '.xlsx'
    wb = openpyxl.load_workbook(filename=nome)
    ws = wb.active
    datas = []
    categoria, subcategoria, origem = '', '', ''
    dados = []
    for cell in range(2, ws.max_column - 3, 4):
        datas.append({"data": ws.cell(row=1, column=cell).value, "coluna": cell})
    for linha in range(1, ws.max_row+1):
        valor_celula = ws.cell(row=linha, column=1).value
        if valor_celula == 'Categ. de despesa':
            origem = 'Despesas'
        elif valor_celula == 'Categ. de rendimento':
            origem = 'Receitas'
        elif str(valor_celula).lower().find('total') > -1:
            continue
        elif valor_celula == None:
            continue
        elif valor_celula[0:2].isnumeric():
            categoria = valor_celula
        elif len(origem) > 0:
                subcategoria = valor_celula
                if subcategoria == 'Todas as outras despesas':
                    categoria = 'OUTROS'
                for data in datas:
                    real = {
                        'origem': origem,
                        'categoria': categoria,
                        'subcategoria': subcategoria,
                        'data': data['data'],
                        'tipo': 'Real',
                        'valor': ws.cell(row=linha, column=data['coluna']).value
                    }
                    orcado = {
                        'origem': origem,
                        'categoria': categoria,
                        'subcategoria': subcategoria,
                        'data': data['data'],
                        'tipo': 'Orçado',
                        'valor': ws.cell(row=linha, column=data['coluna']+1).value
                    }
                    if not real['valor'] == None:
                        dados.append(real)
                    if not orcado['valor'] == None:
                        dados.append(orcado)

    dados.sort(key=lambda x: (x['origem'], x['data']))
    wb.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Real Orçado"
    ws['A1'] = 'Origem'
    ws['B1'] = 'Categoria'
    ws['C1'] = 'Subcategoria'
    ws['D1'] = 'Data'
    ws['E1'] = 'Tipo'
    ws['F1'] = "Valor"
    linha = 2
    for dado in dados:
        ws.cell(row=linha, column=1, value=dado['origem'])
        ws.cell(row=linha, column=2, value=dado['categoria'])
        ws.cell(row=linha, column=3, value=dado['subcategoria'])
        ws.cell(row=linha, column=4, value=dado['data'])
        ws.cell(row=linha, column=5, value=dado['tipo'])
        ws.cell(row=linha, column=6, value=dado['valor'])
        linha += 1

    nome = arguments[3] if arguments[3] else nome
    wb.save(nome)
    print("Automação concluída com sucesso")


def marioluz():
    nome = arguments[2]
    nome += '.xlsx'
    wb = openpyxl.load_workbook(filename=nome)
    ws = wb.active
    columns = [1]
    headerRow = 0
    for line in range(2, ws.max_column - 3, 4):
        if ws.cell(row=line, column=1).value == "(Processo) ID":
            headerRow = line
    data = []
    for column in range(1, ws.max_column+1):
        if ws.cell(row=headerRow, column=column).value == "Status"\
                or ws.cell(row=headerRow, column=column).value == "Processo - Objeto Criminal - Principal"\
                or ws.cell(row=headerRow, column=column).value == "Processo - Centro de Custo Histórico"\
                or ws.cell(row=headerRow, column=column).value == "Data Registrado"\
                or ws.cell(row=headerRow, column=column).value == "Data de encerramento"\
                or ws.cell(row=headerRow, column=column).value == "(Processo) Estado" \
                or ws.cell(row=headerRow, column=column).value == "(Processo) ID":
            obj = {
                "header": ws.cell(row=headerRow, column=column).value,
                "value": []
            }
            for cell in range(headerRow+1, ws.max_row+1):
                obj["value"].append(ws.cell(row=cell, column=column).value)
            data.append(obj)
    wb.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arquivo Modificado"
    coluna = 1
    for dado in data:
        ws.cell(row=1, column=coluna, value=dado['header'])
        coluna += 1

    linha = 2
    coluna = 1
    for dado in data:
        for value in dado["value"]:
            ws.cell(row=linha, column=coluna, value=value)
            linha += 1
        coluna += 1
        linha = 2

    nome = arguments[3]+".xlsx" if len(arguments) > 3 else nome
    wb.save(nome)


if len(arguments) == 1:
    print("Nenhum comando encontrado. Para ajuda com esse script insira o comando -h ou help")
elif arguments[1].lower() == "help" or arguments[1].lower() == "-h":
    print("""Os seguintes comandos estão disponíveis
help, -h              
    ajuda

axpe                
    Automação para axpe com os seguintes argumentos:
        1º caminho de origem (arquivo sem extensão)
        2º caminho de destino (opcional, arquivo sem extensão)

marioluz            
    automação mário luz""")
elif arguments[1].lower() == 'axpe':
    axpe()
elif arguments[1].lower() == 'marioluz':
    marioluz()
else:
    print("Insira um comando válido.")
